import tkinter as tk
from tkinter import messagebox
import math
import numpy as np
import csv
import os
from openpyxl import Workbook, load_workbook
import pandas as pd
from tkinter import simpledialog, filedialog
from tkinter import ttk
import openpyxl
import sys
from PIL import Image, ImageTk
from tkinter import font as tkfont
from openpyxl.styles import Alignment
from decimal import Decimal, ROUND_DOWN

class DistanceCalculator3D:
    def __init__(self, master):
        self.i = 0
        if not os.path.isfile("./data.xlsx"):
            self.create_excel_file()
        
        self.path = "./data.xlsx"
        self.master = master
        master.title("Stability Of Block Boulders")
        self.master.geometry("1520x640")
        
        self.logo_path = resource_path("measurement/libs/img.jpg")
        image = Image.open(self.logo_path)
        resized_image = image.resize((70, 70)) 
        photo = ImageTk.PhotoImage(resized_image)
        
        self.logo_label = tk.Label(master, image=photo)
        self.logo_label.image = photo  
        self.logo_label.grid(row=0, column=0, columnspan=2, sticky="nw")
        
        self.custom_font = tkfont.Font(family="Helvetica", size=12, weight="bold", slant="italic")
        self.text_label = tk.Label(master, text="STABILITY OF BLOCK BOULDERS", font=self.custom_font, fg="blue", bg="#e4eded")
        self.text_label.grid(row=0, column=2, padx=10, pady=10, sticky="nw")
        
        #-------------------------------------------------------------------------

        self.label1 = tk.Label(master, text="Point 1 (x1, y1, h1):")
        self.label1.grid(row=1, column=0, padx=10, pady=5)

        self.entry_x1 = tk.Entry(master)
        self.entry_x1.grid(row=1, column=1, padx=10, pady=5)
        self.entry_x1.insert(tk.END, "0")

        self.entry_y1 = tk.Entry(master)
        self.entry_y1.grid(row=1, column=2, padx=10, pady=5)
        self.entry_y1.insert(tk.END, "0")

        self.entry_z1 = tk.Entry(master)
        self.entry_z1.grid(row=1, column=3, padx=10, pady=5)
        self.entry_z1.insert(tk.END, "0")
        #-------------------------------------------------------------------------
        self.label2 = tk.Label(master, text="Point 2 (x2, y2, h2):")
        self.label2.grid(row=2, column=0, padx=10, pady=5)

        self.entry_x2 = tk.Entry(master)
        self.entry_x2.grid(row=2, column=1, padx=10, pady=5)
        self.entry_x2.insert(tk.END, "0")

        self.entry_y2 = tk.Entry(master)
        self.entry_y2.grid(row=2, column=2, padx=10, pady=5)
        self.entry_y2.insert(tk.END, "0")

        self.entry_z2 = tk.Entry(master)
        self.entry_z2.grid(row=2, column=3, padx=10, pady=5)
        self.entry_z2.insert(tk.END, "0")
        #-------------------------------------------------------------------------
        self.label3 = tk.Label(master, text="Point 3 (x3, y3, h3):")
        self.label3.grid(row=3, column=0, padx=10, pady=5)

        self.entry_x3 = tk.Entry(master)
        self.entry_x3.grid(row=3, column=1, padx=10, pady=5)
        self.entry_x3.insert(tk.END, "0")

        self.entry_y3 = tk.Entry(master)
        self.entry_y3.grid(row=3, column=2, padx=10, pady=5)
        self.entry_y3.insert(tk.END, "0")

        self.entry_z3 = tk.Entry(master)
        self.entry_z3.grid(row=3, column=3, padx=10, pady=5)
        self.entry_z3.insert(tk.END, "0")
        #-------------------------------------------------------------------------
        self.label4 = tk.Label(master, text="Radius (r): ")
        self.label4.grid(row=4, column=0, padx=10, pady=5)
        
        self.entry_r = tk.Entry(master)
        self.entry_r.grid(row=4, column=1, padx=10, pady=5)
        self.entry_r.insert(tk.END, "0")
        
        self.label5 = tk.Label(master, text="The permissible stability")
        self.label5.grid(row=4, column=2, padx=10, pady=5)
        
        self.entry_t = tk.Entry(master)
        self.entry_t.grid(row=4, column=3, padx=10, pady=5)
        self.entry_t.insert(tk.END, "0")
        #-------------------------------------------------------------------------
        self.calculate_button = tk.Button(master, text="Calculate", command=self.rs_FoS)
        self.calculate_button.grid(row=5, column=1, padx=10, pady=5)

        self.cal_from_xlsx = tk.Button(master, text="Exit", command=master.destroy, background="red", foreground="white")
        self.cal_from_xlsx.grid(row=5, column=3, padx=10, pady=5)

        self.compare = tk.Button(master, text="Compare", command=self.compare_action, background="blue", foreground="white")
        self.compare.grid(row=5, column=2, padx=10, pady=5)
        
        self.result_text = tk.Text(master, height=15, width=70)
        self.result_text.grid(row=6, column=0, columnspan=4, padx=40, pady=5)

        self.scrollbar = tk.Scrollbar(master, command=self.result_text.yview)
        # self.scrollbar.place(relx=0.97, rely=0.65, anchor="center", relheight=0.6)
        self.scrollbar.grid(row=6, column=3, sticky="ns")
        self.result_text.config(yscrollcommand=self.scrollbar.set)

        self.create_widgets()

    def create_widgets(self):
        columns = ("STT", "Tên", "X", "Y", "Chiều Dài", "Chiều Rộng", "Chiều Cao", "Loại vật chất tiếp xúc",
                   "X1", "Y1", "H1", "X2", "Y2", "H2", "X3", "Y3", "H3", "R (m)", "FoS", "B'", "Độ ổn định cho phép", "Đánh giá độ ổn định")

        # Tạo Treeview
        self.tree = ttk.Treeview(self.master, columns=columns, show='headings')

        # Đặt tiêu đề cho các cột
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=100)

        # Grid the treeview
        self.tree.grid(row=1, rowspan=10, column=4, columnspan=2, padx=10, pady=10, sticky="nsew")

        # Create horizontal scrollbar
        h_scrollbar = ttk.Scrollbar(self.master, orient=tk.HORIZONTAL, command=self.tree.xview)
        h_scrollbar.grid(row=11, column=4, columnspan=2, sticky="ew")
        self.tree.configure(xscrollcommand=h_scrollbar.set)

        # Create vertical scrollbar
        v_scrollbar = ttk.Scrollbar(self.master, orient=tk.VERTICAL, command=self.tree.yview)
        v_scrollbar.grid(row=1, rowspan=10, column=6, sticky="ns")
        self.tree.configure(yscrollcommand=v_scrollbar.set)

        # self.master.grid_rowconfigure(1, weight=1)
        self.master.grid_columnconfigure(4, weight=1)

        # Add buttons or other widgets if needed
        self.open_excel_button = tk.Button(self.master, text="Open Excel", command=self.open_excel)
        self.open_excel_button.grid(row=12, column=0, padx=10, pady=10)

        self.cal_from_file = tk.Button(self.master, text="Cal From File", command=self.calculate_from_file, background="#32a895")
        self.cal_from_file.grid(row=12, column=3, padx=10, pady=10)
        
    def create_excel_file(self):
        workbook = Workbook()
        sheet = workbook.active

        headers = [
            "STT", "Tên", "Vị Trí Tảng Đá", "", "Kích Thước (m)", "", "", "Loại vật chất tiếp xúc", "VT1", "", "",
            "VT2", "", "", "VT3","", "", "R (m)", "FoS", "B'", "Độ ổn định cho phép", "Đánh giá độ ổn định"
        ]
        
        values = [
            "", "", "X", "Y", "Chiều Dài", "Chiều Rộng", "Chiều Cao", "" ,"X1", "Y1", "H1", "X2", "Y2", "H2", "X3", "Y3", "H3", "", "", "", "", ""
        ]
        
        sheet.append(headers)
        sheet.append(values)
        
        merge_cells = ['A1:A2', 'B1:B2', 'C1:D1', 'E1:G1', 'I1:K1', 'L1:N1', 'O1:Q1', 'H1:H2', 'R1:R2', 'S1:S2', 'T1:T2', 'U1:U2', 'V1:V2', 'W1:W2']
        for merge_cell in merge_cells:
            sheet.merge_cells(merge_cell)
        
        for row in sheet.iter_rows(min_row=1, max_row=2, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.alignment = cell.alignment.copy(wrapText=True)
        
        workbook.save("data.xlsx")  
        
    def calculate_from_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        self.cal_path = file_path
        print(self.cal_path)
        if self.cal_path:
            workbook = openpyxl.load_workbook(self.cal_path)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=3):
                x1, y1, h1 = float(row[8].value), float(row[9].value), float(row[10].value)
                x2, y2, h2 = float(row[11].value), float(row[12].value), float(row[13].value)
                x3, y3, h3 = float(row[14].value), float(row[15].value), float(row[16].value)
                r = float(row[17].value)

                coor1 = (x1, y1, h1)
                coor2 = (x2, y2, h2)
                coor3 = (x3, y3, h3)
                print([i.value for i in row])
                rs, beta = self.cal_FoS(coor1, coor2, coor3, r)
                
                row[18].value = Decimal(rs).quantize(Decimal('0.01'), rounding=ROUND_DOWN) if rs is not None else None
                row[19].value = Decimal(beta).quantize(Decimal('0.01'), rounding=ROUND_DOWN) if beta is not None else None
                   
            
            workbook.save(file_path)
            # self.compare_action(file_path)
            tk.messagebox.showinfo("Thông báo", "Dữ liệu đã được cập nhật và lưu vào file.")
            workbook.close()
            
    def compare_action(self, path = None):
        if path is not None:
            base = path
        else:
            base = self.path
        threshold = float(self.entry_t.get())
        workbook = openpyxl.load_workbook(base)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=3):
            if row[18].value is not None:
                row[20].value = threshold
                row[21].value = "Ổn định" if float(row[18].value) >= threshold else "Không ổn định"
        workbook.save(base)
        self.open_excel(option=1)
        
    def open_excel(self, option = 0):
        if option == 0:
            file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
            self.path = file_path
        else:
            file_path = self.path
        
        # tk.messagebox.showinfo("Thông báo", f"Dữ liệu được thao tác trên file tại đường dẫn {self.path}")
        if file_path:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            self.tree.delete(*self.tree.get_children())

            # header = [cell.value for cell in sheet[1]]

            # self.tree["columns"] = header
            # for col in header:
            #     self.tree.heading(col, text=col)
            #     self.tree.column(col, width=70, stretch=False)
                
            list_value = list(sheet.values)
            for row in list_value[2:]:
                formatted_row = [f"{cell:.2f}" if isinstance(cell, float) else cell for cell in row]
                self.tree.insert("", tk.END, values=formatted_row)
                
        self.tree.bind("<Double-1>", self.on_double_click)
    
    def on_double_click(self, event):
        item = self.tree.selection()[0]
        column = self.tree.identify_column(event.x)
        column_name = self.tree.heading(column)['text']

        current_value = self.tree.item(item, 'values')[self.tree["columns"].index(column_name)]

        entry = tk.Entry(self.master)
        entry.insert(0, current_value)
        entry.place(relx=0.8, rely=0.8, anchor="center")

        def save_changes(event):
            new_value = entry.get()
            print(new_value)
            self.tree.set(item, column=column_name, value=new_value)
            entry.destroy()
            try:
                file_path = self.path
                workbook = openpyxl.load_workbook(file_path)
                sheet = workbook.active
                col_index = self.tree["columns"].index(column_name) + 1
                # row_index = int(item[1:]) + 1  
                for row_idx, row in enumerate(self.tree.get_children(), start=2):
                    if row == item:
                        row_index = row_idx + 1
                        print(row_index)
                        break
                
                cell = sheet.cell(row=row_index, column=col_index)
                cell.value = new_value
                workbook.save(file_path)
                self.update_and_calculate_FoS(item, column_name, new_value)
            except Exception as e:
                print(f"Error saving changes: {e}")

        entry.bind("<Return>", save_changes)
    def update_and_calculate_FoS(self, item, column_name, new_value):
        try:
            file_path = self.path
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            col_index = self.tree["columns"].index(column_name) + 1


            coordinates = [float(self.tree.set(item, column)) for column in ("X1", "Y1", "H1", "X2", "Y2", "H2", "X3", "Y3", "H3")]
            r = float(self.tree.set(item, "R (m)"))


            cell = sheet.cell(row=self.tree.index(item) + 3, column=col_index)
            cell.value = new_value
            print(self.tree.index(item))

            FoS, beta = self.cal_FoS(coordinates[:3], coordinates[3:6], coordinates[6:], r)

            FoS_cell = sheet.cell(row=self.tree.index(item) + 3, column=19)
            FoS_cell.value = Decimal(FoS).quantize(Decimal('0.01'), rounding=ROUND_DOWN) if FoS is not None else None
            
            beta_cell = sheet.cell(row=self.tree.index(item) + 3, column=20)
            beta_cell.value = Decimal(beta).quantize(Decimal('0.01'), rounding=ROUND_DOWN) if beta is not None else None
            print(FoS, beta)
            if sheet.cell(row=self.tree.index(item) + 3, column=21) is not None:
                comment_cell = sheet.cell(row=self.tree.index(item) + 3, column=22)
                comment_cell.value = "Ổn định" if FoS >= float(sheet.cell(row=self.tree.index(item) + 3, column=21).value) else "Không ổn định"
            workbook.save(file_path)
            self.reload()
        except Exception as e:
            print(f"Error updating and calculating FoS: {e}")
            
    def reload(self):

        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook.active
        self.tree.delete(*self.tree.get_children())
            
        list_value = list(sheet.values)
        for row in list_value[2:]:
            print(row)
            self.tree.insert("", tk.END, values=row)
                
        self.tree.bind("<Double-1>", self.on_double_click)
        
    def has_none_value(self, input_tuple):
        return any(value is None for value in input_tuple)
    
    def cal_FoS(self, cordinate1, cordinate2, cordinate3, r):
        if not self.has_none_value(cordinate1) and not self.has_none_value(cordinate2) and not self.has_none_value(cordinate3) and r is not None:
            d = np.sqrt((cordinate1[0] - cordinate2[0])**2 + (cordinate1[1] - cordinate2[1])**2)
            deltah = np.abs(cordinate1[2] - cordinate2[2])
            deltax = np.sqrt(deltah**2 + d**2)
            tan_alpha = np.tan((deltah / d))
            h = np.abs(cordinate3[2] - cordinate2[2])
            
            tan_beta = (deltax - 2*r) / h
            
            FoS = tan_beta / tan_alpha
            return FoS, np.arctan(tan_beta)
        else:
            return None, None
        
    def rs_FoS(self):
        try:
            x1 = float(self.entry_x1.get())
            y1 = float(self.entry_y1.get())
            h1 = float(self.entry_z1.get())
            
            x2 = float(self.entry_x2.get())
            y2 = float(self.entry_y2.get())
            h2 = float(self.entry_z2.get())
            
            x3 = float(self.entry_x3.get())
            y3 = float(self.entry_y3.get())
            h3 = float(self.entry_z3.get())
            
            r = float(self.entry_r.get())
            
            coor1 = (x1, y1, h1)
            coor2 = (x2, y2, h2)
            coor3 = (x3, y3, h3)
            
            rs, beta = self.cal_FoS(coor1, coor2, coor3, r)
            self.result_text.insert(tk.END, f"{coor1}, {coor2}, {coor3}, {r}, FoS : {rs:.3f}, Beta {beta:.3f} \n")
            insert_val = [self.i, "", 0., 0., 0., 0., 0., "",
                   x1, y1, h1, x2, y2, h2, x3, y3, h3, r, rs, beta, float(self.entry_t.get()), ""]
            

            self.open_additional_window(insert_val)
            
        except ValueError:
            messagebox.showerror("Error", "Please enter valid numerical values for the coordinates.")
            
    def open_additional_window(self, val):
        if not hasattr(self, "additional_window") or not self.additional_window.winfo_exists():
            self.additional_window = tk.Toplevel(self.master)
            self.additional_window.title("Confirm Window")

            labels = ["STT", "Tên", "X", "Y", "Chiều Dài", "Chiều Rộng", "Chiều Cao", "Loại vật chất tiếp xúc",
                      "X1", "Y1", "H1", "X2", "Y2", "H2", "X3", "Y3", "H3", "R (m)", "FoS", "B'", "Độ ổn định cho phép", "Đánh giá độ ổn định"]

            self.entry_list = []
            half = len(labels) // 2
            for i, (label_text, entry_text) in enumerate(zip(labels, val)):
                row, col = divmod(i, half)
                label = tk.Label(self.additional_window, text=label_text)
                label.grid(row=row*2, column=col, padx=5, pady=5)
                entry = tk.Entry(self.additional_window)
                entry.grid(row=row*2+1, column=col, padx=5, pady=5)
                entry.insert(tk.END, entry_text)
                self.entry_list.append(entry)

            insert_button = tk.Button(self.additional_window, text="Insert", command=self.insert_value, background="green", foreground="white")
            insert_button.grid(row=(half+1)*2, columnspan=2, column=3, padx=5, pady=5)
            delete_button = tk.Button(self.additional_window, text="Close", command=self.on_additional_window_close, background="red", foreground="white")
            delete_button.grid(row=(half+1)*2, columnspan=2, column=6, padx=5, pady=5)

            self.calculate_button.config(state="disabled")
            self.cal_from_xlsx.config(state="disabled")
            self.additional_window.protocol("WM_DELETE_WINDOW", self.on_additional_window_close)
            self.additional_window.mainloop()

            
    def on_additional_window_close(self):
        self.additional_window.destroy()
        self.calculate_button.config(state="normal")
        self.cal_from_xlsx.config(state="normal")
        
    def insert_value(self):
        try:
            workbook = openpyxl.load_workbook(self.path)
            sheet = workbook.active

            values = [entry.get() for entry in self.entry_list]

            last_row = sheet.max_row + 1
            print(last_row)
            values[0] = last_row - 2
            values[21] = "Ổn định" if values[18] > values[20] else "Không ổn định"
            for col, value in enumerate(values, start=1):
                sheet.cell(row=last_row, column=col).value = value

            # Lưu file Excel
            workbook.save(self.path)

            print("Inserted values into the last row of the Excel file.")
        except Exception as e:
            print(f"Error inserting values: {e}")

        self.on_additional_window_close()
        self.reload()

def resource_path(relative_path):
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)
 
def main():
    root = tk.Tk()
    root.iconbitmap(resource_path('measurement\\icon_app.ico'))
    app = DistanceCalculator3D(root)
    root.mainloop()

if __name__ == "__main__":
    main()
